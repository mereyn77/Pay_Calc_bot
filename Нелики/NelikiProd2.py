# Import data from excel
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
# import xlrd
import pandas as pd
import xlsxwriter
import easygui

filename_nel1=easygui.fileopenbox()
wbn1 = openpyxl.load_workbook(filename_nel1)
sh_nel1 = wbn1.worksheets[0]

# Identifying period (month, year)
period=list(sh_nel1['B5'].value)
month01=period[5]
month02=period[6]
month_num=month01+month02
year01=period[8]
year02=period[9]
year=year01+year02
month_list={'01':'Январь', '02':'Февраль', '03':'Март', '04':'Апрель', '05':'Май', '06':'Июнь', '07':'Июль', '08':'Август', '09':'Сентябрь', '10':'Октябрь', '11':'Ноябрь', '12':'Декабрь'}
month_name=month_list[month_num]
filename_nel2=str('Нелики '+month_name+' 20'+year+' пол. часть.xlsx')

# Print setup
workbook=xlsxwriter.Workbook(filename_nel2)
worksheet1 = workbook.add_worksheet('Детальный')
worksheet1.set_margins(left=0.3, right=0.2, top=0.2, bottom=0.2)
worksheet2 = workbook.add_worksheet('Короткий')
worksheet2.set_margins(left=0.7, right=0.7, top=0.3, bottom=0.2)
workbook.close()

# Saving data to another xlsx
# filename_nel2=
wbn2=openpyxl.load_workbook(filename_nel2)

sh_nel2=wbn2.active
sh_nel2.title='Детальный'


#Registering new style
nelik_style2=NamedStyle(name='nelik_style2')
nelik_style2.font=Font(name='Arial', size=9, bold=False)
border=Side(style='thin', color='000000')
nelik_style2.border=Border(left=border, top=border, right=border, bottom=border)
wbn2.add_named_style(nelik_style2)


list_length=sh_nel1.max_row
range_col=[1, 2, 4, 6]
sh_nel2_row=2
for k in range(8, list_length):
    sh_nel2_col = 1
    if sh_nel1.cell(row=k, column=3).value != None:
        for j in range_col:
            copy_data = sh_nel1.cell(row=k, column=j)
            sh_nel2.cell(row=sh_nel2_row, column=sh_nel2_col).value = copy_data.value
            sh_nel2.cell(row=sh_nel2_row, column=sh_nel2_col).style = 'nelik_style2'
            sh_nel2_col = sh_nel2_col + 1
    elif sh_nel1.cell(row=k, column=3).value == None and sh_nel1.cell(row=k, column=2).value == None:
        for j in range_col:
            copy_data = sh_nel1.cell(row=k, column=j)
            sh_nel2.cell(row=sh_nel2_row, column=sh_nel2_col).value = copy_data.value
            sh_nel2.cell(row=sh_nel2_row, column=sh_nel2_col).style = 'nelik_style2'
            sh_nel2_col = sh_nel2_col + 1
    else:
        continue
    sh_nel2_row = sh_nel2_row + 1

# Column D integers
list_length2=sh_nel2.max_row


# Calculations salespersons
sh_nel2_row=2
for kk in range (2, list_length2):
    sh_nel2.cell(row=kk, column=5).style = 'nelik_style2'
    if sh_nel2.cell(row=kk, column=2).value==None:
        for c_p in range (1, 6):
            cell_paint = wbn2.active.cell(column=c_p, row=kk)
            cell_paint.fill = openpyxl.styles.PatternFill(start_color='ceffce', end_color='ceffce', fill_type='solid')

        copy_dataip=sh_nel2.cell(row=kk, column=1)
        if str(copy_dataip.value).isupper():
            sh_nel2.cell(row=kk, column=3).value=''
            sh_nel2.cell(row=kk, column=4).value=''
            sh_nel2.cell(row=kk, column=5).value=''

        else:
            copy_datakk = sh_nel2.cell(row=kk, column=4)
            sh_nel2.cell(row=kk, column=5).value=copy_datakk.value*6//100
    else:
        continue
    sh_nel2_row = sh_nel2_row + 1

# Calculations managers
dpts=['Вербовская И А санфаянс', 'Вербовская И А керамика', 'Ширяев И.И. Правый берег керамика', 'Рейн И Э керамика', 'Рейн И.Э. БД-2 (керамика)', 'Охотников Л С  Маяковка  монтаж', 'Охотников Л С Ст.Разина монтаж', 'Охотников Л.С.БД4 Энергетиков', 'Рейн И Э сантехника', 'Ширяев И.И. Правый берег кухни, баня', 'Ширяев И.И. Правый берег санфаянс']
sh_nel2_row=2
dpt_style=Font(color='00FF0000', bold=True)
for ik in range (2, list_length2):
    if sh_nel2.cell(row=ik, column=2).value==None:
        if sh_nel2.cell(row=ik, column=1).value in dpts:
            dpt_name=sh_nel2.cell(row=ik, column=1)
            dpt_name.font=dpt_style
            for yell_cell in range(1, 6):
                dpt_name_line=sh_nel2.cell(row=ik, column=yell_cell)
                dpt_name_line.fill=openpyxl.styles.PatternFill(start_color='FFFF00', end_color='ceffce', fill_type='solid')
            copy_dataman=sh_nel2.cell(row=ik, column=4)
            sh_nel2.cell(row=ik, column=5).value=copy_dataman.value*3//100

        else:
            continue
    else:
        continue
    sh_nel2_row = sh_nel2_row + 1




sh_nel2.column_dimensions['A'].width = 8
sh_nel2.column_dimensions['B'].width = 70
sh_nel2.column_dimensions['C'].width = 5
sh_nel2.column_dimensions['D'].width = 8
sh_nel2.column_dimensions['E'].width = 6

# Saving data for Yulz

# sh_nel3 = wbn2.create_sheet('Короткий')

wbn2.active=1
sh_nel3=wbn2.active
sh_nel3['A1'].value=filename_nel2
sh_nel3['A1'].fill=openpyxl.styles.PatternFill(start_color='FFFF00', end_color='ceffce', fill_type='solid')
sh_nel3_row=2
for ii in range(2, list_length2):
    sh_nel3_col = 1
    if sh_nel2.cell(row=ii, column=2).value == None:
        for jj in range(1, 6, 4):
            copy_data2=sh_nel2.cell(row=ii, column=jj)
            sh_nel3.cell(row=sh_nel3_row, column=sh_nel3_col).value=copy_data2.value
            sh_nel3.cell(row=sh_nel3_row, column=sh_nel3_col).style = 'nelik_style2'
            sh_nel3_col = sh_nel3_col + 1
    else:
        continue
    sh_nel3_row=sh_nel3_row+1

sh_nel3_row=2
list_length3=sh_nel3.max_row
for ji in range(2, list_length3):
    # sh_nel3_col=1
    # if sh_nel3.cell(row=sh_nel3_row, column=2).value==None:
    if sh_nel3.cell(row=ji, column=1).value in dpts:
        dpt_name = sh_nel3.cell(row=ji, column=1)
        dpt_name.font = dpt_style
    else:
        continue


# sh_nel3.cell(row=7, column=5).style='nelik_style1'
sh_nel3.column_dimensions['A'].width = 55
sh_nel3.column_dimensions['B'].width = 7
sh_nel3.column_dimensions['C'].width = 1


# Font
col_e=sh_nel3.column_dimensions['E']
col_e.font=Font(name='Arial', size=9, bold=True)


wbn2.save(str(filename_nel2))

