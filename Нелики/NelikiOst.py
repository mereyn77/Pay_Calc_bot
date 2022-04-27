# Import data from excel
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
import os
import pandas as pd
import xlsxwriter
import easygui

filename_nel_all=easygui.fileopenbox()
df=pd.read_excel(filename_nel_all, header=None)
df.to_excel(filename_nel_all+'.xlsx', index=False, header=False)
wb_nall=openpyxl.load_workbook(filename_nel_all+'.xlsx')
sh_nel_all1 = wb_nall.worksheets[0]

# Identifying period (month, year)
period=list(sh_nel_all1['A5'].value)
day01=period[9]
day02=period[10]
day=day01+day02
month01=period[12]
month02=period[13]
month_num=month01+month02
year01=period[15]
year02=period[16]
year=year01+year02
month_list={'01':'Января', '02':'Февраля', '03':'Марта', '04':'Апреля', '05':'Мая', '06':'Июня', '07':'Июля', '08':'Августа', '09':'Сентября', '10':'Октября', '11':'Ноября', '12':'Декабря'}
month_name=month_list[month_num]
fl_nm_nel_all=str(day+' '+month_name+' 20'+year)


#Registering new style
nelik_style3=NamedStyle(name='nelik_style3')
nelik_style3.font=Font(name='Arial', size=9, color='00FF0000', bold=True)
border=Side(style='thin', color='000000')
wb_nall.add_named_style(nelik_style3)


# Coloring Entrepreneurs' cells
list_length=sh_nel_all1.max_row
entrepreneurs=['ИП Кирин А.А.', 'ИП Ли Д.О.', 'ИП Охотников Л.С.', 'ИП Андрюшкин А.В.', 'ИП Охотникова А.В.', 'ИП Полищук П.В.', 'ИП Рейн Ю.И.']
entre_style=Font(color='FF0000', bold=True)
for i in range(10, list_length):
    for j in entrepreneurs:
        if j in sh_nel_all1.cell(row=i, column=1).value:
            for k in range(1, 9):
                entre_name = sh_nel_all1.cell(row=i, column=k)
                entre_name.font=entre_style
                entre_name.fill=openpyxl.styles.PatternFill(start_color='FFFF00', fill_type='solid')
        else:
            continue

# Calculations & Changing None values to 0
for ic in range(11, list_length):
    if sh_nel_all1.cell(row=ic, column=6).value!=None:
        val=int(sh_nel_all1.cell(row=ic, column=6).value)
        qty=int(sh_nel_all1.cell(row=ic, column=7).value)
        sub_total=val*qty
        sh_nel_all1.cell(row=ic, column=8).value=sub_total
    else:
        sh_nel_all1.cell(row=ic, column=8).value=0

for i3 in range(11, list_length):
    for j3 in entrepreneurs:
        if j3 in sh_nel_all1.cell(row=i3, column=1).value:
            agent_subtotals = []
            agent_sum_cell=sh_nel_all1.cell(row=i3, column=8)
        else:
            continue

    yel_cell=sh_nel_all1.cell(row=i3, column=1)
    if yel_cell.font!=Font(color='00FF0000'):
        agent_subtotals.append(sh_nel_all1.cell(row=i3, column=8).value)
        agent_sum_cell.value = sum(agent_subtotals)
    else:
        continue

# Another sheet

wb_nall.create_sheet(index=0, title='Итоговый')
wb_nall.active=0
sh_nel_all2=wb_nall.active
sh_nel_all2.column_dimensions['A'].width = 60
sh_nel_all2.column_dimensions['B'].width = 12

sh_nel_all2_row=2
for i4 in range(11, list_length):
    sh_nel_all2_col = 1
    for j4 in entrepreneurs:
        if j4 in sh_nel_all1.cell(row=i4, column=1).value:
            for k4 in range(1, 9, 7):
                copy_data1=sh_nel_all1.cell(row=i4, column=k4)
                sh_nel_all2.cell(row=sh_nel_all2_row, column=sh_nel_all2_col).value=copy_data1.value
                sh_nel_all2.cell(row=sh_nel_all2_row, column=sh_nel_all2_col).style = 'nelik_style3'
                sh_nel_all2_col = sh_nel_all2_col + 1
        else:
            continue
        sh_nel_all2_row=sh_nel_all2_row+1
commodity=easygui.buttonbox("Выбрать товарное направление",
                            choices=['Керамика', 'Сантехника', 'Самообслужка'])
com=commodity[:4]
wb_nall.save(str('Остатки '+com+' '+fl_nm_nel_all+'.xlsx'))
wb_nall.close()
os.remove(filename_nel_all+".xlsx")

