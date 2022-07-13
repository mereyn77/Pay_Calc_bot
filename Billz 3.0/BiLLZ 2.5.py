# Import data from excel
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
# import xlrd
import pandas as pd
import xlsxwriter
import easygui
from PyQt5 import QtCore, QtGui, QtWidgets

ip_list=["ВИА", "ВИБ", "ВТИ", "ОСО", "ОЛС", "РИЭ", "ИП", "ОЕВ", "КСФ", "КНВ", "ШИИ"]
ip_list.sort()



# Interface of the app ==========================
# class Ui_Dialog(object):
#     def setupUi(self, Dialog):
#         Dialog.setObjectName("Dialog")
#         Dialog.setEnabled(True)
#         Dialog.setMinimumSize(QtCore.QSize(470, 300))
#         Dialog.setMaximumSize(QtCore.QSize(470, 300))
#         font = QtGui.QFont()
#         font.setStyleStrategy(QtGui.QFont.PreferDefault)
#         Dialog.setFont(font)
#         Dialog.setStyleSheet("background-color:#FFCC99")
#         Dialog.setSizeGripEnabled(False)
#         self.checkBox = QtWidgets.QCheckBox(Dialog)
#         self.checkBox.setGeometry(QtCore.QRect(10, 270, 469, 20))
#         self.checkBox.setObjectName("checkBox")
#         self.label = QtWidgets.QLabel(Dialog)
#         self.label.setGeometry(QtCore.QRect(20, 20, 171, 31))
#         font = QtGui.QFont()
#         font.setFamily("Arial Black")
#         font.setPointSize(20)
#         font.setBold(False)
#         font.setItalic(False)
#         font.setUnderline(False)
#         font.setWeight(50)
#         font.setStrikeOut(False)
#         font.setKerning(False)
#         font.setStyleStrategy(QtGui.QFont.NoAntialias)
#         self.label.setFont(font)
#         self.label.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
#         self.label.setAccessibleName("")
#         self.label.setAutoFillBackground(False)
#         self.label.setStyleSheet("color: rgb(255, 255, 255);")
#         self.label.setObjectName("label")
#         self.widget = QtWidgets.QWidget(Dialog)
#         self.widget.setGeometry(QtCore.QRect(90, 170, 108, 76))
#         self.widget.setObjectName("widget")
#         self.verticalLayout = QtWidgets.QVBoxLayout(self.widget)
#         self.verticalLayout.setContentsMargins(0, 0, 0, 0)
#         self.verticalLayout.setSpacing(5)
#         self.verticalLayout.setObjectName("verticalLayout")
#         self.radioButton = QtWidgets.QRadioButton(self.widget)
#         self.radioButton.setEnabled(True)
#         self.radioButton.setChecked(True)
#         self.radioButton.setObjectName("radioButton")
#         self.verticalLayout.addWidget(self.radioButton)
#         self.radioButton_2 = QtWidgets.QRadioButton(self.widget)
#         self.radioButton_2.setObjectName("radioButton_2")
#         self.verticalLayout.addWidget(self.radioButton_2)
#         self.radioButton_3 = QtWidgets.QRadioButton(self.widget)
#         self.radioButton_3.setObjectName("radioButton_3")
#         self.verticalLayout.addWidget(self.radioButton_3)
#         self.widget1 = QtWidgets.QWidget(Dialog)
#         self.widget1.setGeometry(QtCore.QRect(10, 80, 452, 74))
#         self.widget1.setObjectName("widget1")
#         self.gridLayout = QtWidgets.QGridLayout(self.widget1)
#         self.gridLayout.setContentsMargins(0, 0, 0, 0)
#         self.gridLayout.setSpacing(12)
#         self.gridLayout.setObjectName("gridLayout")
#         self.pushButton = QtWidgets.QPushButton(self.widget1)
#         self.pushButton.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton.setObjectName("pushButton")
#         self.gridLayout.addWidget(self.pushButton, 0, 0, 1, 1)
#         self.pushButton_2 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_2.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_2.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_2.setObjectName("pushButton_2")
#         self.gridLayout.addWidget(self.pushButton_2, 0, 1, 1, 1)
#         self.pushButton_3 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_3.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_3.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_3.setObjectName("pushButton_3")
#         self.gridLayout.addWidget(self.pushButton_3, 0, 2, 1, 1)
#         self.pushButton_4 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_4.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_4.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_4.setObjectName("pushButton_4")
#         self.gridLayout.addWidget(self.pushButton_4, 0, 3, 1, 1)
#         self.pushButton_5 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_5.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_5.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_5.setObjectName("pushButton_5")
#         self.gridLayout.addWidget(self.pushButton_5, 0, 4, 1, 1)
#         self.pushButton_6 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_6.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_6.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_6.setObjectName("pushButton_6")
#         self.gridLayout.addWidget(self.pushButton_6, 0, 5, 1, 1)
#         self.pushButton_7 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_7.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_7.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_7.setObjectName("pushButton_7")
#         self.gridLayout.addWidget(self.pushButton_7, 1, 0, 1, 1)
#         self.pushButton_8 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_8.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_8.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_8.setObjectName("pushButton_8")
#         self.gridLayout.addWidget(self.pushButton_8, 1, 1, 1, 1)
#         self.pushButton_9 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_9.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_9.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_9.setObjectName("pushButton_9")
#         self.gridLayout.addWidget(self.pushButton_9, 1, 2, 1, 1)
#         self.pushButton_11 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_11.setMaximumSize(QtCore.QSize(65, 50))
#         self.pushButton_11.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_11.setObjectName("pushButton_11")
#         self.gridLayout.addWidget(self.pushButton_11, 1, 3, 1, 1)
#         self.pushButton_10 = QtWidgets.QPushButton(self.widget1)
#         self.pushButton_10.setMaximumSize(QtCore.QSize(65, 50))
#         font = QtGui.QFont()
#         font.setPointSize(10)
#         font.setBold(False)
#         font.setItalic(False)
#         font.setUnderline(False)
#         font.setWeight(50)
#         self.pushButton_10.setFont(font)
#         self.pushButton_10.setStyleSheet("background-color:#FFFFCC;\n"
# "    font: 10pt;\n"
# "")
#         self.pushButton_10.setObjectName("pushButton_10")
#         self.gridLayout.addWidget(self.pushButton_10, 1, 4, 1, 1)
#
#         self.retranslateUi(Dialog)
#         QtCore.QMetaObject.connectSlotsByName(Dialog)
#         Dialog.setTabOrder(self.pushButton, self.pushButton_2)
#         Dialog.setTabOrder(self.pushButton_2, self.pushButton_3)
#         Dialog.setTabOrder(self.pushButton_3, self.pushButton_4)
#         Dialog.setTabOrder(self.pushButton_4, self.pushButton_5)
#         Dialog.setTabOrder(self.pushButton_5, self.pushButton_6)
#         Dialog.setTabOrder(self.pushButton_6, self.pushButton_7)
#         Dialog.setTabOrder(self.pushButton_7, self.pushButton_8)
#         Dialog.setTabOrder(self.pushButton_8, self.pushButton_9)
#         Dialog.setTabOrder(self.pushButton_9, self.pushButton_11)
#         Dialog.setTabOrder(self.pushButton_11, self.pushButton_10)
#
#         self.add_functions() # Added manually
#
#     def retranslateUi(self, Dialog):
#         _translate = QtCore.QCoreApplication.translate
#         Dialog.setWindowTitle(_translate("Dialog", "BiLLZ 3.0"))
#         self.checkBox.setText(_translate("Dialog", "Удалить исходные файлы после работы программы"))
#         self.label.setText(_translate("Dialog", "BiLLZ 3.0"))
#         self.radioButton.setText(_translate("Dialog", "Оба файла"))
#         self.radioButton_2.setText(_translate("Dialog", "Только прем."))
#         self.radioButton_3.setText(_translate("Dialog", "Только ав."))
#         self.pushButton.setText(_translate("Dialog", "ВИА"))
#         self.pushButton_2.setText(_translate("Dialog", "ВИБ"))
#         self.pushButton_3.setText(_translate("Dialog", "ВТИ"))
#         self.pushButton_4.setText(_translate("Dialog", "КСФ"))
#         self.pushButton_5.setText(_translate("Dialog", "КНВ"))
#         self.pushButton_6.setText(_translate("Dialog", "ИП"))
#         self.pushButton_7.setText(_translate("Dialog", "ОЕВ"))
#         self.pushButton_8.setText(_translate("Dialog", "ОЛС"))
#         self.pushButton_9.setText(_translate("Dialog", "ОСО"))
#         self.pushButton_11.setText(_translate("Dialog", "РИЭ"))
#         self.pushButton_10.setText(_translate("Dialog", "ШИИ"))
#
#     def add_functions(self):
#         self.pushButton.clicked.connect(lambda: self.write_text(self.pushButton.text()))
#         self.pushButton_2.clicked.connect(lambda: self.write_text(self.pushButton_2.text()))
#         self.pushButton_3.clicked.connect(lambda: self.write_text(self.pushButton_3.text()))
#         self.pushButton_4.clicked.connect(lambda: self.write_text(self.pushButton_4.text()))
#         self.pushButton_5.clicked.connect(lambda: self.write_text(self.pushButton_5.text()))
#         self.pushButton_6.clicked.connect(lambda: self.write_text(self.pushButton_6.text()))
#         self.pushButton_7.clicked.connect(lambda: self.write_text(self.pushButton_7.text()))
#         self.pushButton_8.clicked.connect(lambda: self.write_text(self.pushButton_8.text()))
#         self.pushButton_9.clicked.connect(lambda: self.write_text(self.pushButton_9.text()))
#         self.pushButton_10.clicked.connect(lambda: self.write_text(self.pushButton_10.text()))
#         self.pushButton_11.clicked.connect(lambda: self.write_text(self.pushButton_11.text()))
#
#
#
#     def write_text(self, text):
#         global filename
#         filename = text
#         print(filename + 'funcion')
#         return filename
#
#
#
#
#
# if __name__ == "__main__":
#     import sys
#     app = QtWidgets.QApplication(sys.argv)
#     Dialog = QtWidgets.QDialog()
#     ui = Ui_Dialog()
#     ui.setupUi(Dialog)
#     Dialog.show()
#     sys.exit(app.exec_())
#





filename=easygui.buttonbox(msg="Выберите нужного ИП.   Нажмите 'ИП', если необходимо выбрать несколько ИП",
                           choices=(ip_list))
if filename=='ИП':
    filename_1 = easygui.fileopenbox()
    wbb1 = openpyxl.load_workbook(filename_1)
    filename_2 = easygui.fileopenbox()
    wbb2 = openpyxl.load_workbook(filename_2)
    filename_3 = easygui.fileopenbox()
    wbb3 = openpyxl.load_workbook(filename_3)
    sh_1 = wbb1.worksheets[0]
    sh_2 = wbb2.worksheets[0]
    sh_3 = wbb3.worksheets[0]


    # sh_1 = wbb1.worksheets[0]  # To be continued here, not complete.

# If the files don't exist create them
else:
    try:
        df = pd.read_excel(filename+'ав.xls', header=None) # Converts xls to xlsx via pandas
        df.to_excel(filename+'ав.xlsx', index=False, header=False, engine='openpyxl')
        with open(filename+"ав.xlsx") as ofa:
            wbb1 = openpyxl.load_workbook(filename+"ав.xlsx")
    except IOError:
        # If the file does not exist, create it
        wbb1=openpyxl.Workbook(filename+'ав.xlsx')
        wbb1.save(filename+'ав.xlsx')

    try:
        df = pd.read_excel(filename + 'пр.xls', header=None)  # Converts xls to xlsx via pandas
        df.to_excel(filename + 'пр.xlsx', index=False, header=False, engine='openpyxl')
        with open(filename+"пр.xlsx") as ofo:
            wbb2 = openpyxl.load_workbook(filename+"пр.xlsx")
    except IOError:
        wbb2=openpyxl.Workbook(filename+'пр.xlsx')
        wbb2.save(filename+'пр.xlsx')

    wbb1 = openpyxl.load_workbook(filename+'ав.xlsx')
    wbb2 = openpyxl.load_workbook(filename+'пр.xlsx')

    sh_1=wbb1.worksheets[0]
    sh_2=wbb2.worksheets[0]

# Payroll #1 adv
empl_list_1=[]
names_col_length1=sh_1.max_row-18-13
for i in range(19, names_col_length1+19):
    name_value_1=sh_1.cell(row=i, column=4).value
    if name_value_1 not in empl_list_1:
        empl_list_1.append(name_value_1)
    else:
        continue

# Payroll #2 other
empl_list_2=[]
names_col_length2=sh_2.max_row-18-13
for i in range(19, names_col_length2+19):
    name_value_2=sh_2.cell(row=i, column=4).value
    if name_value_2 not in empl_list_2:
        empl_list_2.append(name_value_2)
    else:
        continue

# Payroll #3 other. If there's a third table
if filename == 'ИП':
    empl_list_3=[]
    names_col_length3=sh_3.max_row-18-13
    for i in range(19, names_col_length3+19):
        name_value_3=sh_3.cell(row=i, column=4).value
        if name_value_3 not in empl_list_3:
            empl_list_3.append(name_value_3)
        else:
            continue

# General list of employees
if filename == 'ИП':
    empl_list=list(set(empl_list_1 + empl_list_2 + empl_list_3))
else:
    empl_list=list(set(empl_list_1 + empl_list_2))
empl_list.sort()
payroll1={}
payroll2={}
payroll3={}

# Matching names and sum: two dict.
for j in empl_list:
    for jj in range(19, names_col_length1+19):
        if sh_1.cell(row=jj, column=4).value==j:
            name_sum1=sh_1.cell(row=jj, column=7).value
            payroll1[j]=name_sum1
        else:
            continue

for j in empl_list:
    for jj in range(19, names_col_length2 + 19):
        if sh_2.cell(row=jj, column=4).value == j:
            name_sum2 = sh_2.cell(row=jj, column=7).value
            payroll2[j] = name_sum2
        else:
            continue

if filename == 'ИП':
    for j in empl_list:
        for jj in range(19, names_col_length3 + 19):
            if sh_3.cell(row=jj, column=4).value == j:
                name_sum3 = sh_3.cell(row=jj, column=7).value
                payroll3[j] = name_sum3
            else:
                continue


# Combining the two (three) dict
if filename == 'ИП':
    payroll = {}
    for i in empl_list:
        payroll[i] = payroll1.get(i, 0) + payroll2.get(i, 0) + payroll3.get(i, 0)
    wbb3.close()
else:
    payroll={}
    for i in empl_list:
        payroll[i]=payroll1.get(i, 0) + payroll2.get(i, 0)

# Saving data to another xlsx
wbb2.close()
wbb1.close()

filename_sss=str(filename+'$$$.xlsx')

# Print setup
wbbs=xlsxwriter.Workbook(filename_sss)
worksheet1 = wbbs.add_worksheet('Расклад')
worksheet1.set_margins(left=0.6, right=0.2, top=0.2, bottom=0.2)
wbbs.close()

wbbs=openpyxl.load_workbook(filename_sss)
wsh=wbbs.active

wsh.column_dimensions['A'].width = 4
wsh.column_dimensions['B'].width = 35
wsh.column_dimensions['C'].width = 10
wsh.column_dimensions['D'].width = 7
wsh.column_dimensions['E'].width = 7
wsh.column_dimensions['F'].width = 7
wsh.column_dimensions['G'].width = 7

#Registering new styles
table_style1=NamedStyle(name='table_style1')
table_style1.font=Font(name='Arial', size=9, bold=False)
border=Side(style='thin', color='000000')
table_style1.border=Border(left=border, top=border, right=border, bottom=border)
wbbs.add_named_style(table_style1)

table_style2=NamedStyle(name='table_style2', fill=PatternFill(patternType='solid', fgColor=Color('FFAEB9')), \
                        alignment=Alignment(horizontal='center', vertical='center'))
table_style2.font=Font(name='Arial', size=9, bold=True)
border=Side(style='thin', color='000000')
table_style2.border=Border(left=border, top=border, right=border, bottom=border)
wbbs.add_named_style(table_style2)

table_style3=NamedStyle(name='table_style3', \
                        alignment=Alignment(horizontal='center', vertical='bottom'))
table_style3.font=Font(name='Arial', size=9, bold=False)
border=Side(style='thin', color='000000')
table_style3.border=Border(left=border, top=border, right=border, bottom=border)
wbbs.add_named_style(table_style3)

table_style4=NamedStyle(name='table_style4', \
                        alignment=Alignment(horizontal='right', vertical='bottom'))
wbbs.add_named_style(table_style4)


# Applying styles
for k in range(1, len(empl_list)+2):
    for kk in range(1, 8):
        wsh.cell(row=k, column=kk).style='table_style1'

for k in range(2, len(empl_list)+2):
    for kk in range(4, 8):
        wsh.cell(row=k, column=kk).style='table_style3'

for k in range(1, 8):
    cell_paint = wsh.cell(column=k, row=1)
    cell_paint.style='table_style2'

for k in range(3, len(empl_list)+2, 2):
    for kk in range(1, 8):
        cell_paint=wbbs.active.cell(column=kk, row=k)
        cell_paint.fill=openpyxl.styles.PatternFill(start_color='C1FFC1', end_color='ceffce', fill_type='solid')


# # Banknote counter
wsh['A1'].value='№'
wsh['B1'].value='ФИО'
wsh['C1'].value='Всего'
wsh['D1'].value='5000'
wsh['E1'].value='1000'
wsh['F1'].value='500'
wsh['G1'].value='100'

n=1
row = 2
col = 1
five_thou=[]
one_thou=[]
five_hund=[]
one_hund=[]

for ij in empl_list:
    wsh.cell(row=row, column=col).value=n
    wsh.cell(row=row, column=col+1).value=ij
    wsh.cell(row=row, column=col+2).value=payroll.get(ij)
    wsh.cell(row=row, column=col+3).value=payroll.get(ij)//5000
    five_thou.append(wsh.cell(row=row, column=col+3).value)
    wsh.cell(row=row, column=col+4).value=(payroll.get(ij)%5000)//1000
    one_thou.append(wsh.cell(row=row, column=col+4).value)
    wsh.cell(row=row, column=col+5).value=((payroll.get(ij)%5000)%1000)//500
    five_hund.append(wsh.cell(row=row, column=col+5).value)
    wsh.cell(row=row, column=col+6).value=(((payroll.get(ij)%5000)%1000)%500)//100
    one_hund.append(wsh.cell(row=row, column=col + 6).value)

    row=row+1
    n=n+1

# Totals
row_spacer=wsh.row_dimensions[len(empl_list)+2]
row_spacer.font=Font(name='Arial', size=6, bold=False)

wsh.cell(row=len(empl_list)+3, column=3).value='Итого'
wsh.cell(row=len(empl_list)+3, column=4).value='5000'
wsh.cell(row=len(empl_list)+3, column=5).value='1000'
wsh.cell(row=len(empl_list)+3, column=6).value='500'
wsh.cell(row=len(empl_list)+3, column=7).value='100'

for ki in range(3, 8):
    wsh.cell(row=len(empl_list)+3, column=ki).style='table_style2'
    wsh.cell(row=len(empl_list)+4, column=ki).style='table_style3'

subtotals=list(payroll.values())
total=sum(subtotals)
wsh.cell(row=len(empl_list)+4, column=3).value="{:,}".format(total).replace(',', ' ')
wsh.cell(row=len(empl_list)+4, column=4).value=sum(five_thou)
wsh.cell(row=len(empl_list)+4, column=5).value=sum(one_thou)
wsh.cell(row=len(empl_list)+4, column=6).value=sum(five_hund)
wsh.cell(row=len(empl_list)+4, column=7).value=sum(one_hund)

wbbs.save(str(filename_sss))



