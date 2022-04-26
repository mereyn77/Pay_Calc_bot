# Import data from excel
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
# import xlrd
import pandas as pd
import xlsxwriter
import easygui
import tkinter as tk

ip_list=["ВИА", "ВИБ", "ВТИ", "ОСО", "ОЛС", "РИЭ", "ИП", "ОЕВ", "КСФ", "Г", "ШИИ"]
ip_list.sort()

wnd = tk.Tk()
wnd.title('Billz')
wnd.config(bg='#FFE57E')
wnd.geometry("500x300+550+300")
icon = tk.PhotoImage(file='rub.png')
wnd.iconphoto(False, icon)
btn1 = tk.Button(wnd, text=)


wnd.mainloop()


filename=easygui.buttonbox(msg="Выберите нужного ИП.   Нажмите 'ИП', если необходимо выбрать несколько ИП",
                           choices=(ip_list))
if filename=='ИП':
    filename_1 = easygui.fileopenbox()
    wbb1 = openpyxl.load_workbook(filename_1)
    filename_2 = easygui.fileopenbox()
    wbb2 = openpyxl.load_workbook(filename_2)
    filename_3 = easygui.fileopenbox()
    wbb3 = openpyxl.load_workbook(filename_3)

    # sh_1 = wbb1.worksheets[0]  # To be continued here, not complete.

# If the files don't exist create them
else:
    try:
        df = pd.read_excel(filename+'ав.xls', header=None) # Converts xls to xlsx via pandas
        df.to_excel(filename+'ав.xlsx', index=False, header=False, engine='openpyxl')
        with open(filename+"ав.xlsx") as ofa:
            wbb1 = openpyxl.load_workbook(filename+"ав.xlsx")
    except IOError:
        # If the file does not exist, create it
        wbb1=openpyxl.Workbook(filename+'ав.xlsx')
        wbb1.save(filename+'ав.xlsx')

    try:
        df = pd.read_excel(filename + 'пр.xls', header=None)  # Converts xls to xlsx via pandas
        df.to_excel(filename + 'пр.xlsx', index=False, header=False, engine='openpyxl')
        with open(filename+"пр.xlsx") as ofo:
            wbb2 = openpyxl.load_workbook(filename+"пр.xlsx")
    except IOError:
        wbb2=openpyxl.Workbook(filename+'пр.xlsx')
        wbb2.save(filename+'пр.xlsx')

wbb1 = openpyxl.load_workbook(filename+'ав.xlsx')
wbb2 = openpyxl.load_workbook(filename+'пр.xlsx')
sh_1=wbb1.worksheets[0]
sh_2=wbb2.worksheets[0]

# Payroll #1 adv
empl_list_1=[]
names_col_length1=sh_1.max_row-18-13
for i in range(19, names_col_length1+19):
    name_value_1=sh_1.cell(row=i, column=4).value
    if name_value_1 not in empl_list_1:
        empl_list_1.append(name_value_1)
    else:
        continue

# Payroll #2 other
empl_list_2=[]
names_col_length2=sh_2.max_row-18-13
for i in range(19, names_col_length2+19):
    name_value_2=sh_2.cell(row=i, column=4).value
    if name_value_2 not in empl_list_2:
        empl_list_2.append(name_value_2)
    else:
        continue

# General list of employees
empl_list=list(set(empl_list_1 + empl_list_2))
empl_list.sort()
payroll1={}
payroll2={}

# Matching names and sum: two dict.
for j in empl_list:
    for jj in range(19, names_col_length1+19):
        if sh_1.cell(row=jj, column=4).value==j:
            name_sum1=sh_1.cell(row=jj, column=7).value
            payroll1[j]=name_sum1
        else:
            continue

for j in empl_list:
    for jj in range(19, names_col_length2 + 19):
        if sh_2.cell(row=jj, column=4).value == j:
            name_sum2 = sh_2.cell(row=jj, column=7).value
            payroll2[j] = name_sum2
        else:
            continue

# Combining the two dict
payroll={}
for i in empl_list:
    payroll[i]=payroll1.get(i, 0) + payroll2.get(i, 0)

# Saving data to another xlsx
wbb2.close()
wbb1.close()
filename_sss=str(filename+'$$$.xlsx')

# Print setup
wbbs=xlsxwriter.Workbook(filename_sss)
worksheet1 = wbbs.add_worksheet('Расклад')
worksheet1.set_margins(left=0.6, right=0.2, top=0.2, bottom=0.2)
wbbs.close()

wbbs=openpyxl.load_workbook(filename_sss)
wsh=wbbs.active

wsh.column_dimensions['A'].width = 4
wsh.column_dimensions['B'].width = 35
wsh.column_dimensions['C'].width = 10
wsh.column_dimensions['D'].width = 7
wsh.column_dimensions['E'].width = 7
wsh.column_dimensions['F'].width = 7
wsh.column_dimensions['G'].width = 7

#Registering new styles
table_style1=NamedStyle(name='table_style1')
table_style1.font=Font(name='Arial', size=9, bold=False)
border=Side(style='thin', color='000000')
table_style1.border=Border(left=border, top=border, right=border, bottom=border)
wbbs.add_named_style(table_style1)

table_style2=NamedStyle(name='table_style2', fill=PatternFill(patternType='solid', fgColor=Color('FFAEB9')), \
                        alignment=Alignment(horizontal='center', vertical='center'))
table_style2.font=Font(name='Arial', size=9, bold=True)
border=Side(style='thin', color='000000')
table_style2.border=Border(left=border, top=border, right=border, bottom=border)
wbbs.add_named_style(table_style2)

table_style3=NamedStyle(name='table_style3', \
                        alignment=Alignment(horizontal='center', vertical='bottom'))
table_style3.font=Font(name='Arial', size=9, bold=False)
border=Side(style='thin', color='000000')
table_style3.border=Border(left=border, top=border, right=border, bottom=border)
wbbs.add_named_style(table_style3)

table_style4=NamedStyle(name='table_style4', \
                        alignment=Alignment(horizontal='right', vertical='bottom'))
wbbs.add_named_style(table_style4)


# Applying styles
for k in range(1, len(empl_list)+2):
    for kk in range(1, 8):
        wsh.cell(row=k, column=kk).style='table_style1'

for k in range(2, len(empl_list)+2):
    for kk in range(4, 8):
        wsh.cell(row=k, column=kk).style='table_style3'

for k in range(1, 8):
    cell_paint = wsh.cell(column=k, row=1)
    cell_paint.style='table_style2'

for k in range(3, len(empl_list)+2, 2):
    for kk in range(1, 8):
        cell_paint=wbbs.active.cell(column=kk, row=k)
        cell_paint.fill=openpyxl.styles.PatternFill(start_color='C1FFC1', end_color='ceffce', fill_type='solid')


# # Banknote counter
wsh['A1'].value='№'
wsh['B1'].value='ФИО'
wsh['C1'].value='Всего'
wsh['D1'].value='5000'
wsh['E1'].value='1000'
wsh['F1'].value='500'
wsh['G1'].value='100'

n=1
row = 2
col = 1
five_thou=[]
one_thou=[]
five_hund=[]
one_hund=[]

for ij in empl_list:
    wsh.cell(row=row, column=col).value=n
    wsh.cell(row=row, column=col+1).value=ij
    wsh.cell(row=row, column=col+2).value=payroll.get(ij)
    wsh.cell(row=row, column=col+3).value=payroll.get(ij)//5000
    five_thou.append(wsh.cell(row=row, column=col+3).value)
    wsh.cell(row=row, column=col+4).value=(payroll.get(ij)%5000)//1000
    one_thou.append(wsh.cell(row=row, column=col+4).value)
    wsh.cell(row=row, column=col+5).value=((payroll.get(ij)%5000)%1000)//500
    five_hund.append(wsh.cell(row=row, column=col+5).value)
    wsh.cell(row=row, column=col+6).value=(((payroll.get(ij)%5000)%1000)%500)//100
    one_hund.append(wsh.cell(row=row, column=col + 6).value)

    row=row+1
    n=n+1

# Totals
row_spacer=wsh.row_dimensions[len(empl_list)+2]
row_spacer.font=Font(name='Arial', size=6, bold=False)

wsh.cell(row=len(empl_list)+3, column=3).value='Итого'
wsh.cell(row=len(empl_list)+3, column=4).value='5000'
wsh.cell(row=len(empl_list)+3, column=5).value='1000'
wsh.cell(row=len(empl_list)+3, column=6).value='500'
wsh.cell(row=len(empl_list)+3, column=7).value='100'

for ki in range(3, 8):
    wsh.cell(row=len(empl_list)+3, column=ki).style='table_style2'
    wsh.cell(row=len(empl_list)+4, column=ki).style='table_style3'

subtotals=list(payroll.values())
total=sum(subtotals)
wsh.cell(row=len(empl_list)+4, column=3).value="{:,}".format(total).replace(',', ' ')
wsh.cell(row=len(empl_list)+4, column=4).value=sum(five_thou)
wsh.cell(row=len(empl_list)+4, column=5).value=sum(one_thou)
wsh.cell(row=len(empl_list)+4, column=6).value=sum(five_hund)
wsh.cell(row=len(empl_list)+4, column=7).value=sum(one_hund)

wbbs.save(str(filename_sss))
