# Программа для формирования файла сводов
import os
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Font, GradientFill, Alignment, PatternFill, Color
from openpyxl.writer.excel import save_workbook
from openpyxl import Workbook
import pandas as pd
import xlsxwriter
import easygui
import datetime
from datetime import date

# Date for transformation it into the name of the file
dt = datetime.datetime.now()
dt2 = dt.strftime('%d %b %Y')
month_list={'Jan':'01', 'Feb':'02', 'Mar':'03', 'Apr':'04', 'May':'05', 'Jun':'06', 'Jul':'07',
        'Aug':'08', 'Sep':'09', 'Oct':'10', 'Nov':'11', 'Dec':'12'}
month_num = int(month_list[dt2[3:6]])-1
# dt3 = str(dt2[:3] + month_name + ' ' + dt2[7:13])
dt3 = dt2.replace(dt2[3:9], str(month_num))
dt4 = dt3[3:7]

svodFile = 'Общ'+'_'+ dt4 + '.xlsx'

# Lists of dpts and corresponding sheet names and column titles
wb1 = openpyxl.load_workbook('Свод_образец.xlsx')
ws_file_list = wb1['File_list']

file_list = [] # List of downloaded files
file_list_2 = [] # List of processed files
file_list_length = ws_file_list.max_row
dpt_list = []

for f in range(1, file_list_length+1):
        filename = ws_file_list.cell(row=f, column=2).value
        file_list.append(filename)
        dpt_name = ws_file_list.cell(row=f, column=1).value
        if dpt_name == None:
            continue
        else:
            dpt_list.append(dpt_name)

ws_sample_table = wb1['Sample_table']
col_titles = []
sample_table_length = ws_sample_table.max_column

for c in range(1, sample_table_length+1):
    col_title = ws_sample_table.cell(row=1, column=c).value
    col_titles.append(col_title)

# Create Named Style for proper decoration of the table
named_st1 = NamedStyle(name='tab_bord')
named_st1.font = Font(name='Arial', bold=False, size=8)
border = Side(style='thin', color='000000')
named_st1.border = Border(left=border, top=border, right=border, bottom=border)

# Create an Excel file for consolidating payrolls
wb_Svody = Workbook()
wb_Svody.add_named_style(named_st1)  # Registering the Named Style for further use

# Creating the main file
# Generating column list
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']
gen_col = [g+'1' for g in columns]

for sh in dpt_list:
        ws = wb_Svody.create_sheet(title=sh)
        ws.append(col_titles)
        ws.freeze_panes = 'F2'
        for a in gen_col:
                ws[a].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        for b in columns:
                ws.column_dimensions[b].width=10
        # ws.merge_cells('D1:E1')
        ws.row_dimensions[1].height=36
        ws.column_dimensions['A'].width=5
        ws.column_dimensions['B'].width=25
        ws.column_dimensions['C'].width=13
        ws.column_dimensions['D'].width=5
        ws.column_dimensions['E'].width=7
        ws.column_dimensions['J'].width=9
        ws.column_dimensions['O'].width=12
        ws.column_dimensions['Q'].width=11
        ws.column_dimensions['R'].width=11
        ws.column_dimensions['Y'].width=11
        ws.column_dimensions['Z'].width=11


wb_Svody.remove(wb_Svody['Sheet'])
wb_Svody.save(svodFile)

# File processing
df1 = pd.read_excel('Свод_образец.xlsx')
for f in file_list:
        df2 = pd.read_excel(f)
        df2.to_excel(f + '_1.xlsx')
        wb_f_0 = openpyxl.load_workbook(f + '_1.xlsx')
        ws_f_0 = wb_f_0.active
        c = 3
        for o in range(3,6): # Колонка, которая показывает, сколько часов отработано, пустая, поэтому dataframe не
                # работает. Надо ее заполнить одинаково с образцом, чтобы слияние произошло правильно.
            if ws_f_0.cell(row=1, column=o).value == 'Отработано':
                ws_f_0.cell(row=1, column=o+1).value = 'оч'
                wb_f_0.save(f + '_1.xlsx')
            else:
                continue
        # Concatination of two dataframes: source file and the sample file
        df2 = pd.read_excel(f + '_1.xlsx')
        df3 = pd.concat([df1, df2], axis=0)
        df3.to_excel(f + '_1.xlsx')
        wb_f_1 = openpyxl.load_workbook(f + '_1.xlsx')
        ws_f_1 = wb_f_1.active
        tble_len = ws_f_1.max_row
        ws_f_1.delete_cols(idx=1, amount=1)
        ws_f_1.delete_cols(idx=31, amount=1)
        wb_f_1.save(f + '_1.xlsx')
        file_list_2.append(f + '_1.xlsx')

for f_2 in file_list_2:
        df_n = pd.read_excel(f_2)
        sh_name = f_2[0:3]
        wb_f_2 = openpyxl.load_workbook(svodFile)
        ws_f_2 = wb_f_2[sh_name]
        l = ws_f_2.max_row
        if l == 1:
                r = 2
        else:
                r = l + 3
        with pd.ExcelWriter(svodFile, engine='xlsxwriter') as writer:
            df_n.to_excel(writer, sheet_name=sh_name, startrow=r, startcol=0)
        wb_f_2.save(svodFile)


print(svodFile)
print(file_list_length)
print(file_list)
print(file_list_2)
