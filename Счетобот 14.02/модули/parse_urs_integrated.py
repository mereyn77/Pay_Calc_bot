import pandas as pd
import re
from datetime import datetime
import math

def normalize_department_name(name):
    """Нормализует название отдела для точного сравнения"""
    if pd.isna(name) or not isinstance(name, str):
        return ''
    
    # 1. Удаляем непечатаемые символы (оставляем буквы, цифры, пробелы, дефисы, точки)
    name = ''.join(char for char in name if char.isprintable())
    
    # 2. Заменяем множественные пробелы/табы на один пробел
    name = re.sub(r'\s+', ' ', name)
    
    # 3. Удаляем пробелы в начале и конце
    name = name.strip()
    
    return name

def parse_urs_settings(file_path, sheet_name=0, report_period=None):
    print(f"⚙️  Парсинг файла настроек (новая структура с %): {file_path}")
    print(f"  📅 Отчётный период: {report_period}")
    
    try:
        # Читаем все колонки до R
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, header=None, usecols="A:R")
        print(f"  📊 Размер: {len(df)} строк × {len(df.columns)} колонок")
        
        # ===== 0. ЧТЕНИЕ ЯЧЕЙКИ I2 (НОВЫЙ "ОКЛАД") =====
        print("  🔍 Чтение ячейки I2 (новый 'Оклад')...")
        оклад_I2 = 0
        try:
            # Открываем файл отдельно для чтения ячейки I2
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Ячейка I2 (9-я колонка, 2-я строка)
            cell_value = ws['I2'].value
            if cell_value is not None:
                # Преобразуем в число
                cell_str = str(cell_value).replace(',', '.').replace(' ', '').strip()
                if cell_str and cell_str.lower() not in ['', 'nan', 'none', 'null']:
                    оклад_I2 = float(cell_str)
                    print(f"  ✅ Найден 'Оклад' в ячейке I2: {оклад_I2:,.0f} руб.")
                else:
                    print(f"  ⚠️  Ячейка I2 пустая или содержит: '{cell_value}'")
            else:
                print(f"  ⚠️  Ячейка I2 пустая")
        except Exception as e:
            print(f"  ⚠️  Не удалось прочитать ячейку I2: {e}")
        
        # ===== 1. НАХОДИМ ЗАГОЛОВОК =====
        header_row = None
        for i in range(min(10, len(df))):
            if len(df.columns) > 1:
                col_a = str(df.iloc[i, 0]).lower().strip() if pd.notna(df.iloc[i, 0]) else ""
                col_b = str(df.iloc[i, 1]).lower().strip() if pd.notna(df.iloc[i, 1]) else ""
                
                if 'фирмы и отделы' in col_a and 'отделы' in col_b:
                    header_row = i
                    print(f"  🔍 Найден заголовок новой таблицы в строке {i+1}")
                    print(f"     Колонка A: '{col_a}'")
                    print(f"     Колонка B: '{col_b}'")
                    break
        
        if header_row is None:
            return {
                'success': False,
                'error': 'Не найден заголовок таблицы',
                'exclusions': [],
                'оклад_I2': оклад_I2  # Добавляем даже при ошибке
            }
        
        # ===== 2. ВСЕ ИСКЛЮЧЕНИЯ (колонка A) =====
        exclusions = []
        
        for i in range(header_row + 1, len(df)):
            if len(df.columns) > 0:
                exclusion_cell = str(df.iloc[i, 0]).strip()
                if (exclusion_cell and 
                    exclusion_cell.lower() not in ['nan', 'none', ''] and
                    len(exclusion_cell) > 2 and
                    not exclusion_cell.replace(',', '').replace('.', '').isdigit()):
                    exclusions.append(exclusion_cell)
        
        print(f"  🚫 Всего исключений (A): {len(exclusions)}")
        if exclusions:
            print(f"  📋 Примеры: {exclusions[:5]}")
        
        # ===== 3. ТАБЛИЦА ОТДЕЛОВ (колонка B) И КОЛОНКИ =====
        col_mapping = {}
        for col_idx in range(len(df.columns)):
            cell = str(df.iloc[header_row, col_idx]).lower().strip()
            
            if 'фирмы и отделы' in cell:
                col_mapping['исключения'] = col_idx
            elif 'отделы' in cell:
                col_mapping['отдел_расчет'] = col_idx
            elif 'филиал' in cell:
                col_mapping['филиал'] = col_idx
            elif 'базовая' in cell:
                col_mapping['базовая_часть'] = col_idx  # Переименовываем в 'базовая_часть' ниже
            elif 'средняя' in cell:
                col_mapping['средняя_зп'] = col_idx
            elif 'минимал' in cell:
                col_mapping['минималка'] = col_idx
            elif 'нелик' in cell and '%' not in cell:
                col_mapping['неликвиды'] = col_idx
            elif 'нелик%' in cell or 'нелик %' in cell:
                col_mapping['неликвид_процент'] = col_idx
            elif 'норма час' in cell:
                col_mapping['норма_часов'] = col_idx
            elif 'обычный товар' in cell or 'обычных' in cell:
                col_mapping['коэф_обычных'] = col_idx
            elif 'бонусный товар' in cell or 'бонусных' in cell:
                col_mapping['коэф_бонусных'] = col_idx
            elif 'неликвид' in cell and '%' in cell:
                col_mapping['коэф_неликвидов'] = col_idx
            elif 'опт' in cell and '%' in cell:
                col_mapping['коэф_оптовых'] = col_idx
            elif '1 место' in cell:
                col_mapping['гарантия_1'] = col_idx
            elif '2 место' in cell:
                col_mapping['гарантия_2'] = col_idx
            elif '3 место' in cell:
                col_mapping['гарантия_3'] = col_idx
            elif '4 место' in cell:
                col_mapping['гарантия_4'] = col_idx
            elif '5 место' in cell:
                col_mapping['гарантия_5'] = col_idx
        
        print(f"  📋 Колонки: {col_mapping}")
        
        # ===== 4. ОБРАБОТКА ОТДЕЛОВ (ТОЛЬКО КОЛОНКА B) =====
        departments = {}
        processed = 0
        skipped = 0
        
        for i in range(header_row + 1, len(df)):
            if 'отдел_расчет' not in col_mapping:
                continue
                
            # НОРМАЛИЗАЦИЯ НАЗВАНИЯ ОТДЕЛА
            dept_cell = df.iloc[i, col_mapping['отдел_расчет']]
            dept_name = normalize_department_name(dept_cell)
            
            if not dept_name or dept_name.lower() in ['nan', 'none', '']:
                skipped += 1
                continue

            forbidden_keywords = ['отдел оптовых продаж', 'опт', 'управление', 'склад', 
                                 'хоз.отдел', 'декрет', 'ип', 'водители', 'уволенные']
            if any(keyword in dept_name.lower() for keyword in forbidden_keywords):
                skipped += 1
                continue
            
            if dept_name.lower() in ['офис', 'администрация', 'опт', 'управление', 'склад']:
                skipped += 1
                continue
            
            departments[dept_name] = {'отдел': dept_name}
            processed += 1
        
        # ===== 5. ЗАПОЛНЯЕМ ДАННЫЕ ОТДЕЛОВ =====
        print(f"\n  📊 ЗАПОЛНЕНИЕ НАСТРОЕК ОТДЕЛОВ:")
        for dept_name in departments.keys():
            dept_row_idx = None
            for i in range(header_row + 1, len(df)):
                current_cell = df.iloc[i, col_mapping['отдел_расчет']]
                current_name = normalize_department_name(current_cell)
                if current_name == dept_name:
                    dept_row_idx = i
                    break
            
            if dept_row_idx is None:
                continue
                
            dept_data = departments[dept_name]
            
            # Филиал
            if 'филиал' in col_mapping:
                филиал = normalize_department_name(df.iloc[dept_row_idx, col_mapping['филиал']])
                dept_data['филиал'] = филиал if филиал else 'Не указан'
            
            # БАЗОВАЯ ЧАСТЬ (бывший "Оклад")
            if 'базовая_часть' in col_mapping:
                try:
                    salary = str(df.iloc[dept_row_idx, col_mapping['базовая_часть']]).replace(',', '.').replace(' ', '')
                    dept_data['базовая_часть'] = float(salary) if salary else 0
                except:
                    dept_data['базовая_часть'] = 0
            
            # Добавляем ОКЛАД из ячейки I2 (одинаковый для всех отделов)
            dept_data['оклад'] = оклад_I2
            
            # Средняя ЗП
            if 'средняя_зп' in col_mapping:
                try:
                    avg = str(df.iloc[dept_row_idx, col_mapping['средняя_зп']]).replace(',', '.').replace(' ', '')
                    dept_data['средняя_зп'] = float(avg) if avg else 0
                except:
                    dept_data['средняя_зп'] = 0
            
            # Минималка
            if 'минималка' in col_mapping:
                try:
                    minim = str(df.iloc[dept_row_idx, col_mapping['минималка']]).replace(',', '.').replace(' ', '')
                    dept_data['минималка'] = float(minim) if minim else 0
                except:
                    dept_data['минималка'] = 0
            
            # Неликвиды в котле
            if 'неликвиды' in col_mapping:
                nelik = str(df.iloc[dept_row_idx, col_mapping['неликвиды']]).lower().strip()
                dept_data['неликвиды_в_котле'] = 'да' in nelik
            
            # Процент неликвидов
            if 'неликвид_процент' in col_mapping:
                try:
                    percent = str(df.iloc[dept_row_idx, col_mapping['неликвид_процент']]).replace(',', '.').replace(' ', '')
                    dept_data['неликвид_процент'] = float(percent) if percent else 0
                except:
                    dept_data['неликвид_процент'] = 0
            
            # Коэффициенты товаров
            if 'коэф_обычных' in col_mapping:
                try:
                    coeff = str(df.iloc[dept_row_idx, col_mapping['коэф_обычных']]).replace(',', '.').replace(' ', '')
                    dept_data['коэф_обычных'] = float(coeff) if coeff else 0.0
                except:
                    dept_data['коэф_обычных'] = 0.0

            if 'коэф_бонусных' in col_mapping:
                try:
                    coeff = str(df.iloc[dept_row_idx, col_mapping['коэф_бонусных']]).replace(',', '.').replace(' ', '')
                    dept_data['коэф_бонусных'] = float(coeff) if coeff else 0.0
                except:
                    dept_data['коэф_бонусных'] = 0.0

            if 'коэф_неликвидов' in col_mapping:
                try:
                    coeff = str(df.iloc[dept_row_idx, col_mapping['коэф_неликвидов']]).replace(',', '.').replace(' ', '')
                    dept_data['коэф_неликвидов'] = float(coeff) if coeff else 0.0
                except:
                    dept_data['коэф_неликвидов'] = 0.0

            if 'коэф_оптовых' in col_mapping:
                try:
                    coeff = str(df.iloc[dept_row_idx, col_mapping['коэф_оптовых']]).replace(',', '.').replace(' ', '')
                    dept_data['коэф_оптовых'] = float(coeff) if coeff else 0.0
                except:
                    dept_data['коэф_оптовых'] = 0.0

            # Гарантии мест
            guarantee_columns = [
                ('гарантия_1', col_mapping.get('гарантия_1')),
                ('гарантия_2', col_mapping.get('гарантия_2')), 
                ('гарантия_3', col_mapping.get('гарантия_3')),
                ('гарантия_4', col_mapping.get('гарантия_4')),
                ('гарантия_5', col_mapping.get('гарантия_5'))
            ]
            
            for key, col_idx in guarantee_columns:
                if col_idx is not None:
                    try:
                        cell_value = str(df.iloc[dept_row_idx, col_idx])
                        clean_value = cell_value.replace(',', '.').replace(' ', '').strip()
                        if clean_value and clean_value.lower() not in ['', 'nan', 'none', 'null']:
                            dept_data[key] = float(clean_value)
                        else:
                            dept_data[key] = 0.0
                    except:
                        dept_data[key] = 0.0
                else:
                    dept_data[key] = 0.0
            
            # Норма часов
            if 'норма_часов' in col_mapping:
                norm_type = normalize_department_name(df.iloc[dept_row_idx, col_mapping['норма_часов']]).lower()
                dept_data['тип_нормы'] = norm_type
                
                if norm_type == 'магазин':
                    dept_data['норма_часов'] = None
                elif norm_type == 'офис':
                    dept_data['норма_часов'] = None
                else:
                    dept_data['норма_часов'] = 160
            else:
                dept_data['тип_нормы'] = 'магазин'
                dept_data['норма_часов'] = None
            
            # Проверка заполнения (добавляем оклад из I2)
            print(f"    Отдел '{dept_name[:30]:30}': "
                  f"База={dept_data.get('базовая_часть',0):,.0f}, "
                  f"Оклад(I2)={dept_data.get('оклад',0):,.0f}, "
                  f"Средняя={dept_data.get('средняя_зп',0):,.0f}, "
                  f"Коефы={dept_data.get('коэф_обычных',0)}/{dept_data.get('коэф_бонусных',0)}/{dept_data.get('коэф_неликвидов',0)}/{dept_data.get('коэф_оптовых',0)}, "
                  f"Гарантии={dept_data.get('гарантия_1',0)}/{dept_data.get('гарантия_2',0)}/{dept_data.get('гарантия_3',0)}")
        
        print(f"  ✅ Отделов для расчетов: {processed}")
        
        # ===== 6. СТАТИСТИКА =====
        filials_set = set()
        for dept_data in departments.values():
            filial = dept_data.get('филиал', 'Не указан')
            if filial and filial != 'Не указан':
                filials_set.add(filial)
        
        print(f"\n  📊 ИТОГОВАЯ СТАТИСТИКА:")
        print(f"  • Отделов для расчетов: {len(departments)}")
        print(f"  • Исключений: {len(exclusions)}")
        print(f"  • Уникальных филиалов: {len(filials_set)}")
        print(f"  • Оклад из I2: {оклад_I2:,.0f} руб.")
        
        return {
            'success': True,
            'departments': departments,
            'exclusions': exclusions,
            'оклад_I2': оклад_I2,  # Возвращаем для использования
            'statistics': {
                'departments_count': len(departments),
                'exclusions_count': len(exclusions),
                'unique_filials': len(filials_set),
                'оклад_I2': оклад_I2,
                'processed': processed,
                'skipped': skipped
            }
        }
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'exclusions': [],
            'оклад_I2': 0
        }
